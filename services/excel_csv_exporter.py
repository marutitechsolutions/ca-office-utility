"""
Excel / CSV Exporter Service
Professional Excel and CSV export with formatting, frozen headers, auto-width,
number/date formatting, summary sheet, and exception/review sheet.
"""
import csv
import os
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class OutputMode:
    RAW = "raw"
    CLEAN = "clean"
    PROFESSIONAL = "professional"


class ExcelCSVExporter:
    """Professional Excel and CSV export."""

    # Style constants
    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid") if HAS_OPENPYXL else None
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF") if HAS_OPENPYXL else None
    DATA_FONT = Font(name="Segoe UI", size=10) if HAS_OPENPYXL else None
    EXCEPTION_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") if HAS_OPENPYXL else None
    THIN_BORDER = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), 
                         top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')) if HAS_OPENPYXL else None
    @staticmethod
    def _sanitize_sheet_name(name):
        """Clean sheet names to prevent Excel errors (max 31 chars, no invalid chars)."""
        if not name:
            return "Data"
        clean = re.sub(r'[\\/\?\*\[\]:]', '', str(name)).strip()
        return clean[:31] if clean else "Data"

    @staticmethod
    def export_to_excel(headers, rows, output_path, mode=OutputMode.PROFESSIONAL,
                        exception_rows=None, title="Extracted Data", sheet_name="Data"):
        """
        Export data to a professional Excel file.
        
        Args:
            headers: list of header strings
            rows: list of list[str] data rows
            output_path: path to save the .xlsx file
            mode: OutputMode (raw, clean, professional)
            exception_rows: optional list of rows for exception/review sheet
            title: title for the summary sheet
            sheet_name: name of the main data sheet
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is not installed. Install via: pip install openpyxl")

        wb = Workbook()

        # Main data sheet
        ws = wb.active
        ws.title = ExcelCSVExporter._sanitize_sheet_name(sheet_name)

        if mode == OutputMode.RAW:
            ExcelCSVExporter._write_raw(ws, headers, rows)
        elif mode == OutputMode.CLEAN:
            ExcelCSVExporter._write_clean(ws, headers, rows)
        else:  # professional
            ExcelCSVExporter._write_professional(ws, headers, rows)

        # Exception/Review sheet
        if exception_rows:
            ws_exc = wb.create_sheet("Exceptions")
            exc_headers = headers + ["Reason"]
            ExcelCSVExporter._write_professional(ws_exc, exc_headers,
                                                  [r + ["Low confidence"] for r in exception_rows])

            # Highlight exception rows
            for row_idx in range(2, len(exception_rows) + 2):
                for col_idx in range(1, len(exc_headers) + 1):
                    cell = ws_exc.cell(row=row_idx, column=col_idx)
                    cell.fill = ExcelCSVExporter.EXCEPTION_FILL

        # Summary sheet (professional mode only)
        if mode == OutputMode.PROFESSIONAL:
            ws_sum = wb.create_sheet("Summary", 0)  # Insert at beginning
            ExcelCSVExporter._write_summary(ws_sum, title, len(rows),
                                             len(exception_rows) if exception_rows else 0,
                                             len(headers))

        wb.save(output_path)

    @staticmethod
    def _write_raw(ws, headers, rows):
        """Write data without formatting."""
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)

    @staticmethod
    def _write_clean(ws, headers, rows):
        """Write data with basic formatting."""
        if headers:
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)

        for row in rows:
            ws.append(row)

        # Freeze top row
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_professional(ws, headers, rows):
        """Write data with professional formatting."""
        if headers:
            ws.append(headers)
            try:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = ExcelCSVExporter.HEADER_FONT
                    cell.fill = ExcelCSVExporter.HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if ExcelCSVExporter.THIN_BORDER:
                        cell.border = ExcelCSVExporter.THIN_BORDER
            except Exception:
                pass

        for row_data in rows:
            ws.append(row_data)

        # Format data cells
        for row_idx in range(2, len(rows) + 2):
            num_cols = len(headers) if headers else (len(rows[0]) if rows else 0)
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                try:
                    cell.font = ExcelCSVExporter.DATA_FONT
                    if ExcelCSVExporter.THIN_BORDER:
                        cell.border = ExcelCSVExporter.THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                except Exception:
                    pass

                # Try to format as number
                val = cell.value
                header_name = headers[col_idx-1].lower() if headers and col_idx <= len(headers) else ""
                
                # SENSITIVE TEXT COLUMNS: Skip numeric parsing
                if any(x in header_name for x in ["narration", "details", "ref", "date", "sr", "particular", "page", "range"]):
                    cell.value = str(val) if val is not None else ""
                    cell.data_type = 's'
                    cell.number_format = '@'
                elif isinstance(val, str):
                    # 1. Try Number ONLY for amount-like columns
                    if any(x in header_name for x in ["debit", "credit", "balance", "amount", "tax", "total", "cgst", "sgst", "igst"]):
                        numeric = ExcelCSVExporter._try_parse_number(val)
                        if numeric is not None:
                            cell.value = numeric
                            cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    # 2. Try Date
                    else:
                        date_val = ExcelCSVExporter._try_parse_date(val)
                        if date_val:
                            cell.value = date_val
                            cell.number_format = 'DD/MM/YYYY'
                            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-width columns
        ExcelCSVExporter._auto_width(ws, headers, rows)

    @staticmethod
    def _write_summary(ws, title, data_count, exception_count, col_count):
        """Write a summary sheet."""
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

        rows = [
            ("Report", title),
            ("Generated", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Total Data Rows", data_count),
            ("Exception Rows", exception_count),
            ("Total Columns", col_count),
            ("Software", "CA Office PDF Utility"),
        ]

        for i, (label, value) in enumerate(rows, start=1):
            cell_a = ws.cell(row=i, column=1, value=label)
            cell_b = ws.cell(row=i, column=2, value=value)
            cell_a.font = Font(name="Segoe UI", size=11, bold=True)
            cell_b.font = Font(name="Segoe UI", size=11)

    @staticmethod
    def _auto_width(ws, headers, rows):
        """Set column widths based on content."""
        for col_idx in range(1, (len(headers) if headers else 0) + 1):
            max_len = len(str(headers[col_idx - 1])) if headers else 0
            for row in rows:
                if col_idx - 1 < len(row):
                    cell_len = len(str(row[col_idx - 1]))
                    if cell_len > max_len:
                        max_len = cell_len
            # Cap at reasonable width
            adjusted = min(max_len + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    @staticmethod
    def _try_parse_number(text):
        """Try to parse text as a number."""
        if not text or not isinstance(text, str):
            return None
        # Remove commas, spaces, currency symbols
        # Remove commas, spaces
        cleaned = text.strip().replace(",", "").replace(" ", "")
        # Remove currency symbols and common labels without stripping the decimal point
        cleaned = re.sub(r'(?i)₹|\$|Rs\.?|INR', '', cleaned)
        
        # Handle trailing Dr/Cr or sign
        is_neg = False
        if "-" in cleaned or "Dr" in text.upper():
            is_neg = True
            
        # Keep only digits and decimal point
        cleaned = re.sub(r'[^0-9\.]', '', cleaned)
        try:
            if not cleaned: return None
            val = float(cleaned)
            return -val if is_neg else val
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _try_parse_date(text):
        """Try to parse text as a date object for Excel."""
        if not text or not isinstance(text, str) or len(text) < 8:
            return None
        
        # Common formats in statements
        formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d/%m/%y", "%d-%b-%Y"
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text.strip(), fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def export_to_csv(headers, rows, output_path, encoding="utf-8-sig"):
        """
        Export data to CSV file.
        
        Args:
            headers: list of header strings
            rows: list of list[str] data rows
            output_path: path to save the .csv file
            encoding: file encoding (utf-8-sig for Excel compatibility)
        """
        with open(output_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)
            if headers:
                writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
